import streamlit as st
import pandas as pd
import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from zoneinfo import ZoneInfo
import time
import threading

def calcular_tempo(arquivo_excel):
    """Calcula o tempo restante e excedente na planilha."""
    
    def formatar_tempo(delta):
        total_segundos = int(delta.total_seconds())
        horas, resto = divmod(total_segundos, 3600)
        minutos, segundos = divmod(resto, 60)
        return f"{horas:02d}:{minutos:02d}:{segundos:02d}"
    
    try:
        df = pd.read_excel(arquivo_excel)
    except FileNotFoundError:
        df = pd.DataFrame(columns=['Item', 'Operador', 'Termino', 'Tempo restante', 'Tempo excedente'])
        df.to_excel(arquivo_excel, index=False)
        st.warning(f"Arquivo {arquivo_excel} criado com colunas padrão.")

    if 'Tempo restante' not in df.columns:
        df['Tempo restante'] = ''
    if 'Tempo excedente' not in df.columns:
        df['Tempo excedente'] = ''

    for index, row in df.iterrows():
        if pd.notna(row.get('Termino')):
            termino_str = str(row['Termino'])
            try:
                termino = datetime.datetime.strptime(termino_str, '%d/%m/%Y %H:%M:%S')
            except ValueError:
                try:
                    hora = datetime.datetime.strptime(termino_str, '%H:%M:%S').time()
                    data_atual = datetime.date.today()
                    termino = datetime.datetime.combine(data_atual, hora)
                except ValueError:
                    st.error(f"Formato inválido na linha {index + 2}: {termino_str}")
                    continue

            agora = datetime.datetime.now(ZoneInfo("America/Sao_Paulo"))
            termino = termino.replace(tzinfo=ZoneInfo("America/Sao_Paulo"))

            if termino > agora:
                tempo_restante = termino - agora
                df.at[index, 'Tempo restante'] = formatar_tempo(tempo_restante)
                df.at[index, 'Tempo excedente'] = 'Dentro do tempo'
            else:
                tempo_excedente = agora - termino
                df.at[index, 'Tempo restante'] = 'Expirado'
                df.at[index, 'Tempo excedente'] = formatar_tempo(tempo_excedente)
        else:
            df.at[index, 'Tempo restante'] = ''
            df.at[index, 'Tempo excedente'] = ''

    df.to_excel(arquivo_excel, index=False)
    
    # Aplicar formatação condicional
    wb = load_workbook(arquivo_excel)
    ws = wb.active
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    # Formatar linha inteira
    for row in range(2, len(df) + 2):
        ws.conditional_formatting.add(
            f'A{row}:E{row}',
            CellIsRule(operator='equal', formula=['"Expirado"'], fill=red_fill)
        )
    
    wb.save(arquivo_excel)
    
    return df

def monitorar():
    while st.session_state.monitoring:
        agora = time.time()
        if agora - st.session_state.last_refresh > st.session_state.refresh_rate:
            st.session_state.last_refresh = agora
            st.rerun()
        time.sleep(1)

def main():
    st.title("Monitoramento de Tempo em Tempo Real ⏱️")
    
    # Inicializar estados
    if 'monitoring' not in st.session_state:
        st.session_state.monitoring = False
    if 'last_refresh' not in st.session_state:
        st.session_state.last_refresh = time.time()
    if 'refresh_rate' not in st.session_state:
        st.session_state.refresh_rate = 60

    # Upload de arquivo
    with st.expander("📤 Carregar Planilha", expanded=True):
        uploaded_file = st.file_uploader(
            "Selecione o arquivo monitoramento.xlsx",
            type=["xlsx"],
            accept_multiple_files=False
        )
        
        if uploaded_file is not None:
            with open("monitoramento.xlsx", "wb") as f:
                f.write(uploaded_file.getbuffer())
            st.success("Arquivo carregado com sucesso!")
            st.session_state.monitoring = False

    # Controles na sidebar
    st.sidebar.header("Controles")
    col1, col2 = st.sidebar.columns(2)
    
    with col1:
        if st.button("▶️ Iniciar"):
            if not st.session_state.monitoring:
                st.session_state.monitoring = True
                st.session_state.last_refresh = time.time()
                threading.Thread(target=monitorar, daemon=True).start()
    
    with col2:
        if st.button("⏹️ Parar"):
            st.session_state.monitoring = False

    # Seletor de intervalo
    st.session_state.refresh_rate = st.sidebar.selectbox(
        "Intervalo de atualização:",
        options=[5, 15, 30, 60],
        format_func=lambda x: f"{x//60} minutos" if x >= 60 else f"{x} segundos",
        index=0
    )

    # Exibição dos dados
    if st.session_state.monitoring or uploaded_file is not None:
        try:
            df = calcular_tempo("monitoramento.xlsx")
            
            st.subheader("📊 Dados Atualizados")
            st.dataframe(
                df.style.apply(
                    lambda row: ['background-color: #ff0000; color: white'] * len(row)
                    if row['Tempo restante'] == 'Expirado'
                    else [''] * len(row),
                    axis=1
                )[['Item', 'Operador', 'Termino', 'Tempo restante', 'Tempo excedente']],
                height=600
            )
            
            with open("monitoramento.xlsx", "rb") as file:
                st.download_button(
                    label="⏬ Baixar Planilha Atualizada",
                    data=file,
                    file_name="monitoramento_atualizado.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
        except Exception as e:
            st.error(f"Erro: {str(e)}")
    else:
        st.info("⚠️ Carregue um arquivo e clique em 'Iniciar' para começar o monitoramento")

if __name__ == '__main__':
    main()
